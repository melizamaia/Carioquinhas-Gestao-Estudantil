"""Etapa 7 — Vagas ofertadas reais (OferecimentosEvagas/*.xlsx) + join dos 2 formatos.

Fecha a ponta da armadilha #3 no seu sentido pesado: cruzar `QueryA.unidade`
com as bases de oferecimento, que usam chaves diferentes por tipo de unidade:

    unidade 7 dígitos = PÚBLICA   → zfill(7) casa com `Designacao` (totalalunos)
    unidade 5 dígitos = PARCEIRA  → CRE(2 díg) + últimos 3 do CÓDIGO SGA (Parceiras)

Join errado aqui perde as 350 parceiras (~149 mil linhas) sem erro nenhum.

Lê os .xlsx com openpyxl (read_only=True, data_only=True), como exige a stack.
NUNCA carrega a QueryB. Grão-alvo: 2025.

Saídas:
  data/vagas.json          — vagas ofertadas e matrículas por unidade + cobertura do join
  painel.json (augmentado) — bloco `vagas_ofertadas` com os totais reais

Rodar (após 05_gerar_jsons.py): python etl/07_vagas_oferecimento.py
"""

from __future__ import annotations

import json
from datetime import datetime, timezone

import openpyxl

import config as cfg
from db import connect


def _idx(header_row: list, rotulo: str) -> list[int]:
    """Índices das colunas cujo rótulo == rotulo (rótulos repetem por grupamento)."""
    return [i for i, v in enumerate(header_row) if v == rotulo]


def ler_publicas() -> dict[str, dict]:
    """totalalunos: por Designacao(7 díg) -> {matriculados, turmas}. (Públicas.)"""
    wb = openpyxl.load_workbook(cfg.OFERECIMENTOS_DIR / "totaalunoscreche2025.xlsx",
                                read_only=True, data_only=True)
    ws = wb["Consolidado"]
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()
    header = list(linhas[2])          # linha 2 = rótulos reais
    col_desig = header.index("Designacao")
    cols_aluno = _idx(header, "Aluno")
    cols_turma = _idx(header, "Turma")

    out: dict[str, dict] = {}
    for row in linhas[3:]:
        desig = row[col_desig]
        if desig is None:
            continue
        unidade = str(desig).strip().zfill(7)
        alunos = sum(int(row[i]) for i in cols_aluno if isinstance(row[i], (int, float)))
        turmas = sum(int(row[i]) for i in cols_turma if isinstance(row[i], (int, float)))
        if unidade in out:
            out[unidade]["matriculados"] += alunos
            out[unidade]["turmas"] += turmas
        else:
            out[unidade] = {"matriculados": alunos, "turmas": turmas}
    return out


def ler_parceiras() -> dict[str, dict]:
    """Parceiras MAIO-2025: por chave 5 díg (CRE+últ.3 SGA) -> {vagas_ofertadas, matriculados}."""
    wb = openpyxl.load_workbook(cfg.OFERECIMENTOS_DIR / "Parceiras2025.xlsx",
                                read_only=True, data_only=True)
    ws = wb["MAIO -2025"]
    linhas = list(ws.iter_rows(values_only=True))
    wb.close()
    header = list(linhas[1])          # linha 1 = rótulos reais
    col_cre = header.index("CRE")
    col_sga = header.index("CÓDIGO SGA")
    col_meta_total = header.index("Meta Total")
    cols_aluno = _idx(header, "Aluno")     # matriculados por grupamento
    cols_meta = _idx(header, "Meta")       # capacidade por grupamento

    out: dict[str, dict] = {}
    for row in linhas[2:]:
        cre, sga = row[col_cre], row[col_sga]
        if cre is None or sga is None:
            continue
        chave = f"{int(cre):02d}{str(int(sga))[-3:]}"   # CRE(2) + últimos 3 do SGA
        meta_total = row[col_meta_total] if isinstance(row[col_meta_total], (int, float)) else 0
        vagas = sum(int(row[i]) for i in cols_meta if isinstance(row[i], (int, float)))
        alunos = sum(int(row[i]) for i in cols_aluno if isinstance(row[i], (int, float)))
        out[chave] = {
            "vagas_ofertadas": int(meta_total or vagas),
            "matriculados": alunos,
        }
    return out


def main() -> None:
    con = connect()
    a = cfg.read_csv_gz_sql(cfg.QUERY_A)
    # Unidades distintas de 2025 (recorte por unidade, opcao<=5).
    unidades = con.sql(f"""
        SELECT DISTINCT CAST(unidade AS VARCHAR) AS unidade
        FROM {a} WHERE ano={cfg.ANO_ALVO} AND opcao<={cfg.OPCAO_MAX_VALIDA}
          AND unidade IS NOT NULL
    """).df()["unidade"].tolist()
    con.close()

    pub = ler_publicas()
    par = ler_parceiras()
    print(f"xlsx públicas: {len(pub):,} designações | xlsx parceiras: {len(par):,} chaves")

    por_unidade = []
    cob = {"publica": {"total": 0, "casadas": 0}, "parceira": {"total": 0, "casadas": 0}}
    for u in unidades:
        tipo = cfg.tipo_unidade(u)
        if tipo == "publica":
            cob["publica"]["total"] += 1
            info = pub.get(u)
            if info:
                cob["publica"]["casadas"] += 1
                por_unidade.append({"unidade": u, "tipo": tipo, "fonte": "totalalunos",
                                    "vagas_ofertadas": None,
                                    "matriculados": info["matriculados"],
                                    "turmas": info["turmas"]})
        elif tipo == "parceira":
            cob["parceira"]["total"] += 1
            info = par.get(u)
            if info:
                cob["parceira"]["casadas"] += 1
                por_unidade.append({"unidade": u, "tipo": tipo, "fonte": "parceiras",
                                    "vagas_ofertadas": info["vagas_ofertadas"],
                                    "matriculados": info["matriculados"]})

    def pct(d):
        return round(100 * d["casadas"] / d["total"], 1) if d["total"] else 0.0

    print("\nCOBERTURA DO JOIN (unidades de 2025):")
    print(f"  públicas (7 díg):  {cob['publica']['casadas']}/{cob['publica']['total']} "
          f"({pct(cob['publica'])}%)")
    print(f"  parceiras (5 díg): {cob['parceira']['casadas']}/{cob['parceira']['total']} "
          f"({pct(cob['parceira'])}%)")

    vagas_parceiras = sum(r["vagas_ofertadas"] for r in por_unidade
                          if r["tipo"] == "parceira" and r["vagas_ofertadas"])
    matric_parceiras = sum(r["matriculados"] for r in por_unidade if r["tipo"] == "parceira")
    matric_publicas = sum(r["matriculados"] for r in por_unidade if r["tipo"] == "publica")
    print("\nTOTAIS (unidades da fila 2025):")
    print(f"  vagas ofertadas (parceiras, Meta): {vagas_parceiras:,}")
    print(f"  matriculados parceiras: {matric_parceiras:,} | públicas: {matric_publicas:,}")

    meta = {
        "ano": cfg.ANO_ALVO,
        "gerado_em": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "nota": ("Join dos 2 formatos: pública=Designacao(7); parceira=CRE+últ.3 SGA(5). "
                 "Vagas ofertadas (Meta) só existem para parceiras; para públicas o xlsx "
                 "traz matrículas/turmas, não capacidade ofertada."),
    }
    salvar = cfg.DATA_DIR / "vagas.json"
    salvar.write_text(json.dumps({
        "meta": meta,
        "cobertura_join": {
            "publica": {**cob["publica"], "pct": pct(cob["publica"])},
            "parceira": {**cob["parceira"], "pct": pct(cob["parceira"])},
        },
        "totais": {
            "vagas_ofertadas_parceiras": vagas_parceiras,
            "matriculados_parceiras": matric_parceiras,
            "matriculados_publicas": matric_publicas,
        },
        "por_unidade": sorted(por_unidade, key=lambda x: (x["tipo"], x["unidade"])),
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n  → data/vagas.json ({salvar.stat().st_size/1024:.1f} KB)")

    # Augmenta painel.json com o bloco de vagas ofertadas reais.
    painel_path = cfg.DATA_DIR / "painel.json"
    painel = json.loads(painel_path.read_text(encoding="utf-8"))
    painel["vagas_ofertadas"] = {
        "fonte": "OferecimentosEvagas (xlsx 2025)",
        "vagas_ofertadas_parceiras": vagas_parceiras,
        "matriculados_parceiras": matric_parceiras,
        "matriculados_publicas": matric_publicas,
        "cobertura_join_pct": {"publica": pct(cob["publica"]),
                               "parceira": pct(cob["parceira"])},
        "nota": meta["nota"],
    }
    painel_path.write_text(json.dumps(painel, ensure_ascii=False, indent=2), encoding="utf-8")
    print("  → painel.json augmentado com bloco vagas_ofertadas")
    print("\n[OK] Etapa 7 concluída.")


if __name__ == "__main__":
    main()
